# 将数据导入数据库
import pymysql
from openpyxl import load_workbook
import os

total_file = 'book_info.xlsx'
data_path = 'result/'
con = pymysql.connect(host='localhost', port=3306, database='nbsa', charset='utf8mb4',
                      user='root', password='123456')
sql = 'SELECT book_id FROM book WHERE book_name = %s'
insert_sql = 'INSERT INTO comment (book_id,comment_content,comment_emotion,comment_emotion_value) VALUES ' \
             '(%s, %s, %s, %s)'
data = []


def test():
    c1 = con.cursor()
    c1.execute(sql, '1Q84 BOOK 1')
    book_id = c1.fetchone()[0]
    print(f'book_id: {book_id}')
    wb_1 = load_workbook(data_path + '1Q84 BOOK 1.xlsx')
    ws1 = wb_1.active
    for row1 in range(1, ws1.max_row + 1):
        text = str(ws1['A' + str(row1)].value)
        if len(text) > 400:
            continue
        data.append((int(book_id), text, int(ws1['B' + str(row1)].value),
                     float(ws1['C' + str(row1)].value)))
    print(f'数据加载完毕')
    c1.close()
    c2 = con.cursor()
    c2.executemany(insert_sql, data)
    con.commit()
    c2.close()
    con.close()


def op_book_emotion():
    data_list = []
    sql_e1 = 'select count(1) from comment where comment_emotion = %s and book_id = %s'
    cursor = con.cursor()
    cursor.execute('select book_id from book')
    book_id_list = cursor.fetchall()
    for book in book_id_list:
        book_id = book[0]
        cursor.execute(sql_e1, (1, book_id))
        pos_num = cursor.fetchone()[0]
        cursor.execute(sql_e1, (0, book_id))
        neg_num = cursor.fetchone()[0]
        print(f'{book_id}:(pos_num:{pos_num},neg_num:{neg_num})')
        if int(pos_num) == 0:
            continue
        data_list.append(
            (int(book_id), int(pos_num) + int(neg_num), int(pos_num), int(neg_num),
             0.0 if neg_num == 0 else float('%.6f' % (pos_num / neg_num))))
    cursor.close()
    c2 = con.cursor()
    sql_e2 = 'insert into book_emotion (book_id,comment_count,pos_comment_count,neg_comment_count,emotion_rate)' \
             'values (%s,%s,%s,%s,%s)'
    c2.executemany(sql_e2,data_list)
    con.commit()
    c2.close()
    con.close()


if __name__ == '__main__':
    op_book_emotion()
    # test()
    # files = os.listdir(data_path)
    # cursor = con.cursor()
    #
    # for file in files:
    #     cursor.execute(sql, file.replace('.xlsx', ''))
    #     book_id = cursor.fetchone()[0]
    #     print(f'{file} 的book_id: {book_id}')
    #     wb = load_workbook(data_path + file)
    #     ws = wb.active
    #     for row in range(1, ws.max_row + 1):
    #         text = str(ws['A' + str(row)].value)
    #         if len(text) > 400:
    #             continue
    #         data.append((int(book_id), text, int(ws['B'+str(row)].value),
    #                      float(ws['C'+str(row)].value)))
    #     print(f'{file} 数据加载完毕')
    #
    # cursor.close()
    # print('全部数据加载完毕')
    #
    # cursor_update = con.cursor()
    # cursor_update.executemany(insert_sql, data)
    # con.commit()
    # cursor_update.close()
    # con.close()

    # wb = load_workbook(total_file)
    # ws = wb.active
    #
    # for row in range(2, ws.max_row + 1):
    #     if ws['E'+str(row)].value is not None:
    #         continue
    #     book_data.append((str(ws['A'+str(row)].value), '暂无信息',
    #                       int(str(ws['D'+str(row)].value).replace('人评价', '')), str(ws['B'+str(row)].value)))
    # print('数据加载完成')
    # print('数据长度 {}'.format(len(book_data)))
    # cursor = book_con.cursor()
    # cursor.executemany(book_sql, book_data)
    # book_con.commit()
    # cursor.close()
    # book_con.close()
