import os
from openpyxl import load_workbook
import random
from openpyxl import Workbook
import numpy as np

target_dir = u'cleaned_data/'
save_dir = u'comments/'
save_num = 0
comments = []
temp_list = []


# 存储数据
def save_one():
    print(f'列表中还有{len(comments)}条数据')
    wb1 = Workbook()
    ws1 = wb1.active
    global temp_list, save_num
    print(type(temp_list))
    for comment in temp_list:
        ws1.append([comment])

    wb1.save(save_dir + 'comments' + str(save_num) + '.xlsx')
    temp_list = np.array([])
    save_num += 1


if __name__ == '__main__':
    save_num = len(os.listdir(save_dir))
    print(f"comment已存数量{save_num}")
    if save_num is 0:
        save_num += 1

    files = os.listdir(target_dir)
    for file in files:
        print(f'开始加载{file}的数据')
        wb = load_workbook(target_dir + file)
        ws = wb.active

        # 排除已经处理完的数据
        if ws['C1'].value == 'done':
            continue
        num = ws.max_row
        print(f'{file}有{num-1}条评论')
        for row in range(2, num + 1):
            comments.append(ws['B' + str(row)].value)
        wb.save(target_dir + file)

        print(f'列表中已有{len(comments)}条数据')

    print("数据加载成功")
    comments = np.array(comments)
    print(type(comments))

    while len(comments) > 0:
        if len(comments) <= 1048570:
            temp_list = comments
            save_one()
            break
        else:
            index1 = np.random.choice(comments.shape[0], size=1048570, replace=False)
            index2 = np.arange(comments.shape[0])
            temp_list = comments[index1]
            save_one()
            comments = comments[np.delete(index2, index1)]

    print("over")
