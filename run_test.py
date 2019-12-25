import jieba

from analyzer import SentimentAnalyzer
from openpyxl import load_workbook
import os
from openpyxl import Workbook

model_path = 'models/model_1225_01.pkl'
test_file_path = 'train_data/train_data.xlsx'
userdict_path = 'userdict.txt'
stopword_path = 'stop_words.txt'
data_path = 'cleaned_data/'
result_path = 'result/'
# corpus_path = './data/review.csv'

# 训练和测试的样本
comments = []
tendency = []
# 预测的结果
predict = []
# 预测器
analyzer = SentimentAnalyzer(model_path=model_path, stopword_path=stopword_path, userdict_path=userdict_path)


# text = '因为听说是游记,我耐着性子看完,最后,还是失望了...'
#
# pos = analyzer.analyze(text=text)


# 加载数据
def init_data(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        comments.append(str(ws['A' + str(row)].value).replace('\n', ''))
        if ws['B' + str(row)].value is None:
            print(row)
        tendency.append(int(ws['B' + str(row)].value))


# 加载停用词
def load_stopwords(path):
    stop_words = []
    with open(path, encoding='UTF-8') as words:
        stop_words.extend([w.strip for w in words.readlines()])
    return stop_words


# 给评论分词
def comment_to_text(comment):
    stop_words = load_stopwords('stop_words.txt')
    comment = jieba.cut(comment)
    all_stop_words = set(stop_words)
    comment_word = [w for w in comment if w not in all_stop_words]
    return comment_word


# 分析数据
def analyze_all_data():
    files = os.listdir(data_path)
    for file in files:
        wb = load_workbook(data_path + file)
        ws = wb.active
        wb_file = Workbook()
        ws_file = wb_file.active
        for row in range(2, ws.max_row + 1):
            val = ws['B' + str(row)].value
            if val is None:
                continue
            text = str(val)
            pos = analyzer.analyze(text=text)
            res = 1 if pos >= 0.88 else 0
            ws_file.append([text, res, pos])
        wb_file.save(result_path + file)
        print(f'{file} 数据分析完毕')


# 测试准确率
def test_predict_rate():
    init_data(test_file_path)
    # 正确数量
    correct = 0
    # 正面数量
    pos_num = 0
    # 负面数量
    neg_num = 0
    # 正面正确数量
    pos_correct = 0
    # 负面正确数量
    neg_correct = 0

    for i in range(0, len(comments)):
        if tendency[i] == 1:
            pos_num += 1
        else:
            neg_num += 1

        text = comments[i]
        pos = analyzer.analyze(text=text)
        res = 1 if pos >= 0.88 else 0
        if res == tendency[i]:
            if tendency[i] == 1:
                pos_correct += 1
            else:
                neg_correct += 1

            correct += 1
        else:
            print(text)
            print(f'情感值:{pos} 预测值:{res} 真实值:{tendency[i]}')

    print('预测正确数量 : {}'.format(str(correct)))
    print('正确率 : {}'.format(str(correct/len(comments))))
    print(f'正面情感正确率 :{pos_correct/pos_num}')
    print(f'负面情感正确率 :{neg_correct/neg_num}')


if __name__ == '__main__':
    # jieba.load_userdict('./userdict.txt')
    test_predict_rate()





