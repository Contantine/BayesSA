from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import random
import os
import daili
import mogu
import datetime

# url
url = 'http://book.douban.com/top250?start='
url_proxy = 'http://www.xicidaili.com/nn/'
# 获取到的代理ip地址
ip_list = []
# proxy,当前代理
proxy_now = {}
# headers
headers_proxy = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/'
                  '53.0.2785.143 Safari/537.36'
}
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/'
                  '78.0.3904.87 Safari/537.36',
    'Host': 'book.douban.com'
}
my_headers = [
    "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 "
    "Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 "
    "Safari/537.75.14",
    "Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)",
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11',
    'Opera/9.25 (Windows NT 5.1; U; en)',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)',
    'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)',
    'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.8.0.12) Gecko/20070731 Ubuntu/dapper-security Firefox/1.5.0.12',
    'Lynx/2.8.5rel.1 libwww-FM/2.14 SSL-MM/1.4.1 GNUTLS/1.2.9',
    "Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.7 (KHTML, like Gecko) Ubuntu/11.04 Chromium/16.0.912.77 "
    "Chrome/16.0.912.77 Safari/535.7",
    "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:10.0) Gecko/20100101 Firefox/10.0 "

]
# 书籍简要信息存储excel文件
file_name = u'book_info.xlsx'
# 书籍简要信息集合
result = []
# 爬取书本位置
book_index = 0
# 书籍地址
exm_url = u''
# 书籍名
exm_name = u''
# 短评数据集合
short_comments = []
# 当前爬取的页数
current_page = 0
# 结束循环的标记
over_flag = False
# 代理是否有效
proxy_flag = False
# 上一次获取代理API的时间
last_time_get_api = datetime.datetime.now()
# 连续没有跳页标签最大值
none_page = 0


# 爬取内容
def spider_douban():
    start = 0
    while start <= 225:
        res1 = requests.get(url + str(start), headers=headers, allow_redirects=False)

        soup = BeautifulSoup(res1.text, 'lxml')
        books = soup.find_all('td', {"valign": "top"})

        for book in books:
            book_info = book.find('a', {"title": True})
            # 去除有些为None的结果
            if book_info is not None:
                # 获得书名和书本的地址
                book_url = book_info['href']
                book_title = book_info['title']
            else:
                continue
            # 获得评分和评价人数
            book_rate = book.find('span', class_='rating_nums').get_text()
            book_comment_nums = book.find('span', class_='pl').get_text().replace(" ", "").replace("\n", "").replace(
                "(", "") \
                .replace(")", "")
            # print("书名:", book_title, " 地址:", book_url, " 评分:", book_rate, " 评价人数:", book_comment_nums)
            result.append({"书名": book_title, "地址": book_url, "评分": book_rate, "评价人数": book_comment_nums})

        start += 25


# 将爬取内容临时存入excel
def file_save(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["书名", "地址", "评分", "评价人数"])
    for book in result:
        ws.append([book['书名'], book['地址'], book['评分'], book['评价人数']])

    wb.save(file_name)


# 有关代理池的函数开始
# 获得代理池
def get_ip_pool():
    return mogu.get_ips()


# 从代理池中取出一个ip
def get_ip():
    if len(ip_list) > 0:
        return {'https': ip_list[0], 'http': ip_list[0]}
    else:
        print('列表中没有元素')


# 删除一个ip
def remove_ip(ip):
    if ip in ip_list:
        ip_list.remove(ip)
        print(f'删除无效ip:{ip},代理池中还有{len(ip_list)}个ip')
    else:
        print(f'ip:{ip}不在代理池中，请检查')


# 获得代理池的大小
def list_len():
    return len(ip_list)


# 有关代理池的函数结束


# 根据书籍url爬取短评
def spider_comments(book_name, book_url, suffix, page_now):
    global current_page, proxy_now, proxy_flag
    # 当proxy失效，重新获取proxy
    if not proxy_flag:
        proxy_now = get_ip()
        print(f'获得ip:{proxy_now}')
    proxy = proxy_now
    try:
        s = requests.session()
        s.keep_alive = False
        res = s.get(book_url + suffix, proxies=proxy,
                    headers={"user-agent": random.choice(my_headers), 'Host': 'book.douban.com'}, allow_redirects=False,
                    timeout=10)
        if res.status_code == 200:
            # 证明该代理有效
            proxy_flag = True

            # print('获得网页成功')
            soup = BeautifulSoup(res.text, 'lxml')
            comments = soup.findAll('li', class_='comment-item')

            # 获得该页中的短评
            for comment in comments:
                user_name = comment.find('span', class_='comment-info').a.get_text()
                comment_short = comment.find('span', class_='short').get_text()
                # print("用户名:", user_name, " 短评:", comment_short)
                short_comments.append({"用户名": user_name, "短评": comment_short})

            # 需要清理数组，防止内存溢出;及时存储数据，防止程序出错而没有保存数据
            if page_now % 100 == 0:
                save_comments(book_name, short_comments, current_page)
                short_comments.clear()

            # 获取页面中的跳页标签
            comment_paginator = soup.find('ul', class_='comment-paginator')
            if comment_paginator is not None and comment_paginator.find('li') is not None:
                global none_page
                none_page = 0
                # 获取“后一页”标签
                next_page = soup.find('ul', class_='comment-paginator').findAll('li')[2].a
            else:
                print(f'第{page_now}页没有跳页标签,爬取下一页内容')
                none_page += 1
                current_page += 1
                if none_page >= 500:
                    print("连续五百页没有跳页标签，更换爬取目标书籍")
                    save_comments(book_name, short_comments, page_now)
                    short_comments.clear()
                    next_book(book_index)
                    global over_flag
                    over_flag = True
                    none_page = 0
                    return
                # 超过一百页就换一个代理ip
                if none_page % 100 is 0:
                    print("连续一百页没有跳页标签，更换ip的地址")
                    deal_with_exp(book_name, page_now, proxy_now['https'])
                return
            # 是否有后一页
            if next_page is not None:
                print(f'第{page_now}页爬取成功!')
                current_page += 1
            else:
                save_comments(book_name, short_comments, page_now)
                short_comments.clear()
                next_book(book_index)

                # 结束标记置为true
                over_flag = True
                print(f"{book_name} 短评爬取完毕,共{page_now}页")

        # 页面被重定向
        elif res.status_code == 302 or res.status_code == 301:
            print(f"爬取第{page_now}页发生重定向")
            print(res.headers['Location'])
            deal_with_exp(book_name, page_now, proxy_now['https'])
        else:
            print(f"返回码:{res.status_code}")
            deal_with_exp(book_name, page_now, proxy_now['https'])
            print("爬取发生错误!")
    except Exception as error:
        print(f"连接失败！进行再次尝试,错误信息:{error}")
        print(f"报错行数:第{error.__traceback__.tb_lineno}行")
        deal_with_exp(book_name, page_now, proxy_now['https'])


# 爬取出现错误或被重定向时的操作
def deal_with_exp(name, page, ip):
    # 存储已经爬到的数据
    save_comments(name, short_comments, page - 1)
    # 清除已经存储的数据
    short_comments.clear()
    # 该代理ip已经无效
    global proxy_flag, proxy_now
    proxy_flag = False
    proxy_now = {}
    # 删除出错的代理ip
    remove_ip(ip)
    # 代理池大小为空时需要重新获取ip
    if list_len() <= 0:
        print("代理ip池为空，正在等待重新获取")
        global last_time_get_api
        # 当前时间
        time_now = datetime.datetime.now()
        # 设置获取时间间隔，防止过于频繁的获取而导致的拒绝访问
        time_interval = (time_now - last_time_get_api).seconds
        if time_interval <= 20:
            time.sleep(20 - time_interval)
        last_time_get_api = datetime.datetime.now()
        global ip_list
        ip_list = get_ip_pool()
        if len(ip_list) is not 0:
            print('代理ip列表获取成功')
            print(ip_list)
    # 重新爬取
    # spider_comments(exm_name, exm_url, f'hot?p={page}', page)


# 保存爬取的短评
def save_comments(book_name, comments, page=-1):
    if len(comments) == 0:
        print('没有爬到数据')
        return
    # 判断文件是否存在
    file_path = f'short_comments\{book_name}.xlsx'
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

    ws = wb.active
    if ws['A1'] is None:
        ws.append(['用户名', '短评'])
    for short in comments:
        ws.append([short['用户名'], short['短评']])

    # 记录爬到的页数
    if page is not -1:
        ws['C1'] = page

    wb.save(file_path)
    if page == -1:
        print('数据持久化成功！')
    else:
        print(f"到第{page}页持久化成功!本次持久化共有{len(comments)}条数据")


# 获得之前已经爬取到的页数，防止爬取重复数据
def get_page(book_name):
    file_path = f'short_comments\{book_name}.xlsx'
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        if ws['C1'] is None:
            return 0
        else:
            return int(ws['C1'].value)
    else:
        return 0


# 获取之前已经爬取完成的书籍
def get_book():
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        print("还没有书籍简要信息，无法爬取短评数据")
        return False
    ws = wb.active

    global book_index
    # 获取要爬取的书籍坐标
    book_index_value = ws['E1'].value
    if book_index_value is None:
        return False
    else:
        # is_done = ws['E'+str(int(book_index_value) + 1)].value
        # while is_done is not "待爬" and book_index_value <= 251:
        #     book_index_value += book_index_value + 1
        #     is_done = ws['E'+str(int(book_index_value) + 1)].value
        book_index = int(book_index_value)
        global exm_url, exm_name
        # 获取要爬取的书籍名字和地址
        if ws['B' + str(book_index + 1)].value is None:
            return False
        exm_url = ws['B' + str(book_index + 1)].value + 'comments/'
        exm_name = ws['A' + str(book_index + 1)].value
        if exm_name is None:
            return False
        else:
            return True


# 更新已经爬完的书籍
def next_book(book_index_now):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        print("还没有书籍简要信息，无法记录,请检查")
        return False

    ws = wb.active
    ws['E1'] = book_index_now + 1
    global book_index, exm_url, exm_name
    wb.save(file_name)
    book_index = book_index_now + 1
    return True


if __name__ == '__main__':

    requests.adapters.DEFAULT_RETRIES = 5
    ip_list = get_ip_pool()
    if len(ip_list) is not 0:
        print('代理ip列表获取成功')
        print(ip_list)
    while get_book():
        print(f"书名:{exm_name},url:{exm_url},第 {book_index} 本书")
        current_page = get_page(exm_name) + 1
        over_flag = False
        none_page = 0
        while over_flag is not True:
            spider_comments(exm_name, exm_url, f'hot?p={current_page}', current_page)
    print("书籍全部爬取完成")
