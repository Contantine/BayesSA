import numpy as np
import pandas as pd
from pandas import DataFrame, Series
import os
from openpyxl import Workbook
from openpyxl import load_workbook

file_name = u'book_info.xlsx'
dir_prefix = u'short_comments/'
target_dir = u'cleaned_data/'
train_data = []

if __name__ == '__main__':
    files = os.listdir(target_dir)
    wb2 = Workbook()
    ws2 = wb2.active
    print(type(ws2))
    for file in files:
        train_data = []
        wb = load_workbook(target_dir + file)
        ws = wb.active
        # ws['B1'] = '用户名'
        # ws['C1'] = '短评内容'
        # ws['D1'] = ''
        # ws.delete_cols(1)
        # wb.save(target_dir + file)
        # print(f'{file}修改完成')
        num = ws.max_row
        for row in range(2, num + 1):
            train_data.append(ws['B' + str(row)].value)
        temp_list = np.array(train_data)
        index = np.random.choice(temp_list.shape[0], size=100, replace=False)
        test_data = temp_list[index]
        for data in test_data:
            ws2.append([data])

        print(f'{file} 选择数据完毕')

    wb2.save('./train_data/train_comments.xlsx')