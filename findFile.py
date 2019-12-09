import os
from openpyxl import Workbook
from openpyxl import load_workbook

file_path = u'short_comments'
book_info = u'book_info.xlsx'
done_files = []
all_books = []
if __name__ == '__main__':
    # 获得爬取完成的书籍名字
    files = os.listdir(file_path)
    for file in files:
        file = file.replace(".xlsx", "")
        done_files.append(file)
    # print(done_files)

    wb = load_workbook(book_info)
    ws = wb.active
    for i in range(2, 251):
        book_name = ws['A' + str(i)].value
        all_books.append(book_name)

    for book in done_files:
        if book in all_books:
            all_books.remove(book)

    print(f'还有{len(all_books)}本书没有爬取')
    print(all_books)
