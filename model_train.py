import pickle
from openpyxl import load_workbook
import jieba
from sklearn import metrics
from sklearn.feature_extraction.text import CountVectorizer, TfidfTransformer, TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline
import numpy as np
import os

file_path = u'train_data/train_data.xlsx'
# 训练和测试数据
comments = []
tendency = []
# 训练模型导出路径
model_export_path = 'model.pkl'


# 加载数据
def init_data():
    wb = load_workbook(file_path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        comments.append(str(ws['A' + str(row)].value).replace('\n', ''))
        if ws['B' + str(row)].value is None:
            print(row)
        tendency.append(int(ws['B' + str(row)].value))


# 加载停用词
def load_stopwords(path):
    stop_words = []
    with open(path, encoding='UTF-8') as words:
        stop_words.extend([i.strip for i in words.readlines()])
    return stop_words


# 给评论分词
def comment_to_text(comment):
    stop_words = load_stopwords('stop_words.txt')
    comment = jieba.cut(comment)
    all_stop_words = set(stop_words)
    comment_word = [w for w in comment if w not in all_stop_words]
    return comment_word


if __name__ == '__main__':
    jieba.load_userdict('./userdict.txt')

    init_data()
    n = len(comments) // 5
    # 训练集
    train_comment_list, train_tendency_list = comments[n:], tendency[n:]
    # 测试集
    test_comment_list, test_tendency_list = comments[:n], tendency[:n]
    print('训练集数量： {}'.format(str(len(train_comment_list))))
    print('测试集数量： {}'.format(str(len(test_comment_list))))

    comment_train = [' '.join(comment_to_text(comment)) for comment in train_comment_list]
    tendency_train = train_tendency_list

    comment_test = [' '.join(comment_to_text(comment)) for comment in test_comment_list]
    tendency_test = test_tendency_list

    # 词频特征向量
    vec = TfidfVectorizer(max_df=0.8, min_df=3)

    # tfidf_vec = TfidfVectorizer()
    #
    #
    # def MNB_Classifier():
    #     return Pipeline([
    #         ('count_vec', CountVectorizer()),
    #         ('mnb', MultinomialNB())
    #     ])
    #
    #
    # mnbc_clf = MNB_Classifier()
    #
    # # 进行训练
    # mnbc_clf.fit(comment_train, tendency_train)
    #
    # # 测试集准确率
    # print('测试集准确率： {}'.format(mnbc_clf.score(comment_test, tendency_test)))

    # tf = TfidfTransformer()

    # 先转换成词频矩阵，再计算TFIDF值
    tf_idf = vec.fit_transform(comment_train)
    # 朴素贝叶斯中的多项式分类器
    clf = MultinomialNB(alpha=0.001).fit(tf_idf, tendency_train)

    test_vec = TfidfVectorizer(max_df=0.5, min_df=3, vocabulary=vec.vocabulary_)
    test_tf = test_vec.fit_transform(comment_test)
    predict_res = clf.predict(test_tf)

    # 0.9238657551274083
    print('测试集准确率： {}'.format(metrics.accuracy_score(tendency_test, predict_res)))
    # 将模型保存pickle文件
    with open(model_export_path, 'wb') as file:
        d = {
            "clf": clf,
            "vectorizer": vec
            # "tfidftransformer": tf,
        }
        pickle.dump(d, file)

    print('训练完成')
